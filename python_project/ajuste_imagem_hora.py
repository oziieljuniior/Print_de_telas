import os
import time
from openpyxl import Workbook

#C:\\Users\\Riallen\\Documents\\Print_de_telas\\data2\\odds
path = '/home/oziel/Documentos/Personal_project/Aviator/Print_de_telas/data2/odds'
t = len(os.listdir(path))
i = 0

odds_hora = []

while i < t:
    name = "/" + str(i) + "img.png"
    #print(name)
    path_odds = path + name
    #print(path_odds)
    file_info = os.stat(path_odds)
    creation_time = time.ctime(file_info.st_ctime)
    print("Hora de criação:", creation_time)
    odds_hora.append(creation_time)

    i += 1

wb = Workbook()
ws = wb.active
wb.title = "Hora_Odds"

ws['A1'] = 'Time_Odds'
indece_excel = 2
i = 0
while i < t:
    ab = 'A' + str(indece_excel)
    ws[ab] = odds_hora[i]
    print(odds_hora[i])

    i += 1
    indece_excel += 1
    

#C:\\Users\\Riallen\\Documents\\Print_de_telas\\python_project 
wb.save("/home/oziel/Documentos/Personal_project/Aviator/Print_de_telas/python_project/odds_hora.xlsx ")
